import re
import os
import base64
import argparse
import time
import json
import itertools
import base64
import requests
import traceback

import cv2
import numpy as np
import easyocr
import openpyxl
import openai

from difflib import SequenceMatcher
from contextlib import contextmanager

os.environ["OPENCV_FFMPEG_READ_ATTEMPTS"] = "100000"


class GPTWrapper:
    def __init__(self, model_name):
        self.result = None
        self.model_name = model_name
        self.client = openai.OpenAI(
            api_key=conf["openai_key"]
        )

    def query(self, query_str, query_img):
        try:
            self.result = self.client.chat.completions.create(
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": query_str,
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpg;base64,{query_img}"
                                }
                            }
                        ]
                    }
                ],
                model=self.model_name,
            )
        except Exception as e:
            print(traceback.format_exc())
            self.result = None

    def get_result(self):
        if self.result:
            raw = self.result.dict()["choices"][0]["message"]["content"]
            raw_fixed = raw.replace('```json\n', '')
            raw_fixed = raw_fixed.replace('\n```', '')
            return raw_fixed

        return None


@contextmanager
def open_cv_video(filepath):
    cap = cv2.VideoCapture(filepath)
    try:
        yield cap
    finally:
        cap.release()


def match_names(res):
    #return len([x for x in res if x.strip().lower() in player_names_set])
    return len([x for x in res if x])

def match_scores(res):
    return len([int(x) for x in res if x.isdigit()])

def match_track(res):
    if not res: return -1
    rez = " ".join(res)
    return max([SequenceMatcher(None, rez, x).ratio() for x in track_names])

def matched_candidates(img, x1, y1, x2, y2, match_func):
    image = img[y1:y2, x1:x2]
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    #ocr = conf["ocr"].readtext(image, detail=0)
    #https://github.com/JaidedAI/EasyOCR/issues/341#issuecomment-2044059424
    scale_factor = 2
    upscaled = cv2.resize(image, None, fx=scale_factor, fy=scale_factor, interpolation=cv2.INTER_LINEAR)
    blur = cv2.blur(upscaled, (5, 5))
    #ocr = conf["ocr"].readtext(blur , detail=0, text_threshold=0.3) #, allowlist='0123456789'
    ocr = conf["ocr"].readtext(blur , detail=0)

    return image, match_func(ocr), ocr

def analyze_frames(video_capture, conf):
    nr_candidates = 0
    video_frame_nr = -1
    #fps = int(video_capture.get(cv2.CAP_PROP_FPS))
    fps = video_capture.get(cv2.CAP_PROP_FPS)
    #sample_rate = conf["movie_sample_rate"] or fps

    next_score_ms = 99999999999999999999999
    next_track_ms = -1   # Note: we assume to start with checking for tracks
    last_score_ms = -1
    last_track_ms = -1
    track_detected = False
    score_detected = False
    score_best = -1
    track_best = -1

    while video_capture.isOpened():
        ret, frame = video_capture.read()
        if not ret:
            break

        video_frame_nr += 1
        msecs = int(video_capture.get(cv2.CAP_PROP_POS_MSEC)) #video_frame_nr//fps
        secs = int(msecs / 1000)

        if conf["movie_start"] and secs < conf["movie_start"]:
            continue
        elif conf["movie_start"] and secs == conf["movie_start"]:
            print(f"MOVIE START ({conf['movie_start']} secs) reached (framenr {video_frame_nr})")
            conf["movie_start"] = -conf["movie_start"]
        if conf["movie_end"] and secs >= conf["movie_end"]:
            print(f"MOVIE END ({conf['movie_end']} secs) reached  (framenr {video_frame_nr})")
            break

        if msecs > next_score_ms:

            if score_detected and msecs - last_score_ms > 10 * 1000:
                # after 10 secs we give up and continue to check for scores
                next_track_ms = msecs + 10 * 1000
                next_score_ms = msecs + 100 * 1000
                score_detected = False
                score_best = -1
                #print("10secs timeout, next!!!")

            img_cropped_scores, match_count_scores, ocr = matched_candidates(frame, conf["aabb_scores_x1"], conf["aabb_scores_y1"], conf["aabb_scores_x2"], conf["aabb_scores_y2"], match_scores)
            if match_count_scores >= conf["threshold_scores"]:

                img_cropped_names, match_count_names, ocr = matched_candidates(frame, conf["aabb_names_x1"], conf["aabb_names_y1"], conf["aabb_names_x2"], conf["aabb_names_y2"], match_names)
                if match_count_names >= conf["threshold_names"]:

                    scoree = min(match_count_scores,12) + min(match_count_names,12)
                    if (scoree) < score_best:
                        continue

                    score_best = scoree

                    if msecs - last_score_ms > 90*1000:
                        # Found a new first score candidate
                        last_score_ms = msecs
                        #print("first hit, set new lastscore ts")

                    #print(f"writing: {conf['frame_dir']}/{video_frame_nr // fps}.jpg")
                    cv2.imwrite(f"{conf['frame_dir']}/{msecs}.jpg", frame)
                    cv2.imwrite(f"{conf['frame_dir']}/{msecs}-names.jpg", img_cropped_names)
                    cv2.imwrite(f"{conf['frame_dir']}/{msecs}-scores.jpg", img_cropped_scores)
                    print("CANDIDATE SCORE FRAME FOUND: ", video_frame_nr, match_count_names, match_count_scores, f"{conf['frame_dir']}/{video_frame_nr // fps}.jpg")
                    nr_candidates+=1

                    next_score_ms = msecs + 1 * 1000
                    score_detected = True

        elif msecs > next_track_ms:

            if track_detected and msecs - last_track_ms > 10 * 1000:
                # after 10 secs we give up and continue to check for scores
                next_track_ms = msecs + 100 * 1000
                next_score_ms = msecs + 90 * 1000
                track_detected = False
                track_best = -1

            # check if we find a track
            img_cropped_track, match_count_track, ocr = matched_candidates(frame, conf["aabb_track_x1"], conf["aabb_track_y1"], conf["aabb_track_x2"], conf["aabb_track_y2"], match_track)

            if match_count_track >= conf["threshold_track"]:

                if match_count_track <= track_best:
                    continue

                track_best = min(match_count_track,12)

                rez = " ".join(ocr)
                best_match=0
                trackname = ''
                for trax in track_names:
                    sc = SequenceMatcher(None, rez, trax).ratio()
                    if sc > best_match:
                        best_match = sc
                        trackname = trax

                track_name = base64.b64encode(trackname.encode()).decode('ascii')
                ff=f"{conf['frame_dir']}/tracks/{msecs}-track-{track_name}.jpg"
                cv2.imwrite(f"{conf['frame_dir']}/tracks/{msecs}.jpg", frame)
                cv2.imwrite(ff, img_cropped_track)

                track_detected = True
                print("CANDIDATE TRACK FRAME FOUND: ", video_frame_nr, trackname, match_count_track, f"{conf['frame_dir']}/tracks/{msecs}.jpg")

                if msecs - last_track_ms > 90 * 1000:
                    # Found a new first track candidate
                    last_track_ms = msecs

                next_track_ms = msecs + 1 * 1000


    print(f"ANALYZED {video_frame_nr} frames ({secs} seconds of video)")
    return nr_candidates


# Main function
if __name__ == "__main__":

    ap = argparse.ArgumentParser()
    ap.add_argument("--action", required=True,
                    help="what action to perform (test, extract, parse, run) (run executes extract & parse in one go)")

    ap.add_argument("--input_aabb_frame", required=False, default=None,
                    help="An image frame to test AABB coordinates & thresholds against")
    ap.add_argument("--input_movie", required=True, help="the movie file to parse")
    ap.add_argument("--input_sheet", required=True,
                    help="the excel sheet to act as input (first sheet should contain player names on row A, scores on row B and NR sheets on row C1 (so we can continue scores in for stiched movie clips))")
    ap.add_argument("--output_frames_dir", required=False, default="None",
                    help="the directory for the output of the text detection pass")
    ap.add_argument("--output_sheet", required=False, default="None",
                    help="the directory for the output of the text detection pass")
    ap.add_argument("--sample_rate", type=int, required=False, default=None,
                    help="only capture every X frame (defaults to every 100th frame)")
    ap.add_argument("--openai_key", required=True, default=None, help="offset (in seconds) when to start capturing")
    ap.add_argument("--movie_start", type=int, required=False, default=None, help="offset (in seconds) when to start capturing")
    ap.add_argument("--movie_end", type=int, required=False, default=None, help="offset (in seconds) when to end capturing")
    ap.add_argument("--aabb_names", required=True, help="the dimensions of the aabb to check for names: x1,y1,x2,y2")
    ap.add_argument("--aabb_scores", required=True, help="the dimensions of the aabb to check for scores: x1,y1,x2,y2")
    ap.add_argument("--aabb_track", required=True,
                    help="the dimensions of the aabb to check for the track name: x1,y1,x2,y2")
    ap.add_argument("--threshold_names", type=float, required=False, default=12,
                    help="the threshold to use for the expected nr of AABBS to detect for a frame with player names")
    ap.add_argument("--threshold_scores", type=float, required=False, default=12,
                    help="the threshold to use for the expected nr of AABBS to detect for a frame with scores")
    ap.add_argument("--threshold_track", type=float, required=False, default=0.6,
                    help="the threshold to use for the expected nr of AABBS to detect for a frame with the track name")

    params = vars(ap.parse_args())

    aabb_names = [int(x) for x in params["aabb_names"].split(",")[0:4]]
    aabb_scores = [int(x) for x in params["aabb_scores"].split(",")[0:4]]
    aabb_track = [int(x) for x in params["aabb_track"].split(",")[0:4]]

    action = params["action"]

    conf = {
        "ocr": easyocr.Reader(['en']),
        "movie_file": params["input_movie"],
        "movie_sample_rate": params["sample_rate"],
        "movie_start": params["movie_start"],
        "movie_end": params["movie_end"],
        "frame_dir": params["output_frames_dir"],

        "openai_key": params["openai_key"],

        "threshold_names": params["threshold_names"],
        "threshold_scores": params["threshold_scores"],
        "threshold_track": params["threshold_track"],

        "aabb_names_x1": aabb_names[0],
        "aabb_names_y1": aabb_names[1],
        "aabb_names_x2": aabb_names[2],
        "aabb_names_y2": aabb_names[3],

        "aabb_scores_x1": aabb_scores[0],
        "aabb_scores_y1": aabb_scores[1],
        "aabb_scores_x2": aabb_scores[2],
        "aabb_scores_y2": aabb_scores[3],

        "aabb_track_x1": aabb_track[0],
        "aabb_track_y1": aabb_track[1],
        "aabb_track_x2": aabb_track[2],
        "aabb_track_y2": aabb_track[3],
    }

    player_names = []
    player_scores = {}
    pos_scores = [15, 12, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]
    # scraped from: https://mkwrs.com/mk8dx/
    track_names = ["Mario Kart Stadium", "Water Park", "Sweet Sweet Canyon", "Thwomp Ruins", "Mario Circuit",
                   "Toad Harbor", "Twisted Mansion", "Shy Guy Falls", "Sunshine Airport", "Dolphin Shoals",
                   "Electrodrome", "Mount Wario", "Cloudtop Cruise", "Bone-Dry Dunes", "Bowser's Castle",
                   "Rainbow Road", "Moo Moo Meadows", "Mario Circuit", "Cheep Cheep Beach", "Toad's Turnpike",
                   "Dry Dry Desert", "Donut Plains 3", "Royal Raceway", "DK Jungle", "Wario Stadium", "Sherbet Land",
                   "Music Park", "Yoshi Valley", "Tick-Tock Clock", "Piranha Plant Slide", "Grumble Volcano",
                   "Rainbow Road", "Yoshi Circuit", "Excitebike Arena", "Dragon Driftway", "Mute City",
                   "Wario's Gold Mine", "Rainbow Road", "Ice Ice Outpost", "Hyrule Circuit", "Baby Park", "Cheese Land",
                   "Wild Woods", "Animal Crossing", "Neo Bowser City", "Ribbon Road", "Super Bell Subway", "Big Blue",
                   "Tour Paris Promenade", "Toad Circuit", "Choco Mountain", "Wii Coconut Mall", "Tour Tokyo Blur",
                   "Shroom Ridge", "Sky Garden", "Tour Ninja Hideaway", "Tour New York Minute", "Mario Circuit 3",
                   "Kalimari Desert", "Waluigi Pinball", "Tour Sydney Sprint", "Snow Land", "Mushroom Gorge",
                   "Sky-High Sundae", "Tour London Loop", "Boo Lake", "Rock Rock Mountain", "Maple Treeway",
                   "Tour Berlin Byways", "Peach Gardens", "Tour Merry Mountain", "Rainbow Road", "Tour Amsterdam Drift",
                   "Riverside Park", "DK Summit", "Yoshi's Island", "Tour Bangkok Rush", "Mario Circuit",
                   "Waluigi Stadium", "Tour Singapore Speedway", "Tour Athens Dash", "Daisy Cruiser",
                   "Moonview Highway", "Squeaky Clean Sprint", "Tour Los Angeles Laps", "Sunset Wilds", "Koopa Cape",
                   "Tour Vancouver Velocity", "Tour Rome Avanti", "DK Mountain", "Daisy Circuit",
                   "Tour Piranha Plant Cove", "Tour Madrid Drive", "Rosalina's Ice World", "Bowser Castle 3",
                   "Rainbow Road"]

    # Parse initial input (player names, scores
    path = params["input_sheet"]
    workbook = openpyxl.load_workbook(path)
    config_sheet = workbook["config"]  # workbook.active
    player_rec = config_sheet.cell(row=2, column=1)
    score_rec = config_sheet.cell(row=2, column=2)
    nr_prev_games_parsed = int(config_sheet.cell(row=2, column=3).value)
    if nr_prev_games_parsed > 0: nr_prev_games_parsed+=1
    cx = 2
    while player_rec.value and cx < 100:
        player_name = player_rec.value.strip().lower()
        player_score = int(score_rec.value)
        print(f"Loaded {cx - 1} - {player_name}: {player_score}")
        player_names.append(player_name)
        player_scores[player_name] = player_score
        cx += 1
        player_rec = config_sheet.cell(row=cx, column=1)
        score_rec = config_sheet.cell(row=cx, column=2)
    print("\n")

    player_names_set = set(player_names)

    if action == "run" or action == "extract":
        if not os.path.exists(conf["frame_dir"]):
            os.makedirs(conf["frame_dir"])
            os.makedirs(conf["frame_dir"]+"/tracks")
        else:
            if not os.path.exists(conf["frame_dir"] + "/tracks"):
                os.makedirs(conf["frame_dir"] + "/tracks")

            try:
                root, _, files = os.walk(conf['frame_dir']).__next__()
                if len(files) > 0:
                    raise Exception("output_frames_dir is not empty!")
            except StopIteration:
                pass
            try:
                root, _, files = os.walk(conf['frame_dir']+"/tracks").__next__()
                if len(files) > 0:
                    raise Exception("output_frames_dir/tracks is not empty!")
            except StopIteration:
                pass

    if action == "test":

        for root, _, files in os.walk(conf['frame_dir']):
            #files.sort(key=lambda f: int(re.sub('\D', '', f[:f.find('-')])))
            for filename in files:
                prefix_names = filename.find("-names")
                prefix_scores = filename.find("-scores")
                # only consider full screen capture frames
                if prefix_names > -1 or prefix_scores > -1:
                    print(f"skip {filename}")
                    continue

                filenamef = f"{conf['frame_dir']}/{filename}"
                frame = cv2.imread(filenamef)

                print(filename)
                img, score, ocr = matched_candidates(frame, conf["aabb_track_x1"], conf["aabb_track_y1"], conf["aabb_track_x2"], conf["aabb_track_y2"], match_track)
                print("track", score, ocr)
                cv2.imwrite(f"{conf['frame_dir']}/{filename}".replace(".png", "-track.png"), img)
                img, score, ocr = matched_candidates(frame, conf["aabb_names_x1"], conf["aabb_names_y1"], conf["aabb_names_x2"], conf["aabb_names_y2"], match_names)
                print("names", score, ocr)
                cv2.imwrite(f"{conf['frame_dir']}/{filename}".replace(".png","-names.png"), img)
                img, score, ocr = matched_candidates(frame, conf["aabb_scores_x1"], conf["aabb_scores_y1"], conf["aabb_scores_x2"], conf["aabb_scores_y2"], match_scores)
                print("scores", score, ocr)
                cv2.imwrite(f"{conf['frame_dir']}/{filename}".replace(".png","-scores.png"), img)
                print(f"{filenamef}: {score}\n\n")

            exit(1)

    if action == "extract" or action == "run":
        # As a preprocessing step, we jump through frames and apply text detection using the EAST text model
        start = time.time()
        with open_cv_video(conf["movie_file"]) as cap:
            candidate_frames = analyze_frames(cap, conf)
        end = time.time()
        print("\nFound {} candidates (took {:.2f} seconds)\n".format(candidate_frames, end - start))

    if action == "scores" or action == "run":
        llm = GPTWrapper(model_name="gpt-4o-2024-05-13")

        screen_idx = 0
        timestamp = 0
        timestamp_prev = -1000
        for root, _, files in os.walk(conf['frame_dir']):
            files.sort(key=lambda f: int(re.sub('\D', '', f[:f.find('-')])))
            for filename in files:
                raw_filename = filename
                filename = f"{conf['frame_dir']}/{filename}"
                prefix_idx = filename.find("-names.jpg")
                if prefix_idx == -1:
                    #print(f"skip file {filename}")
                    continue

                timestamp = int(raw_filename[:raw_filename.find("-")])

                if timestamp - timestamp_prev < 20:
                    print(f"skipping duplicate: {filename}")
                    continue

                print(f"parsing image {filename}")
                with open(filename, 'rb') as image_file:
                    image_data = base64.b64encode(image_file.read()).decode('utf-8')

                query_to_send = "Can you return an array formatted as a json response with the names you see in the provided list? Please return only json in your response formatted like this: [name1,name2,name3,...]"  # Update with your query
                llm.query(query_to_send, image_data)
                try:
                    detected = json.loads(llm.get_result())
                except Exception as e:
                    print(traceback.format_exc())
                    import pdb;pdb.set_trace()

                if len(detected) != len(player_names):
                    print(f"skipping capture: {filename} detected {len(detected)} expected {len(player_names)}")
                    continue

                timestamp_prev = timestamp

                # normalize player names and match based on similarity (gptvision miss interprets some names)
                for dx in range(len(detected)):
                    det = detected[dx].strip().lower()
                    best_match = -1
                    for tx in player_names:
                        score = SequenceMatcher(None, det, tx).ratio()
                        if score > best_match:
                            detected[dx] = tx
                            best_match = score

                #if set(detected) == player_names_set:
                print(f"adding score screen {nr_prev_games_parsed+screen_idx+1}")

                worksheet = workbook.create_sheet(f"Ronde {nr_prev_games_parsed+screen_idx+1}", nr_prev_games_parsed+screen_idx+1)
                worksheet.column_dimensions["A"].width = 30
                worksheet.column_dimensions["B"].width = 10

                for rank, pname in enumerate(detected):
                    score = pos_scores[rank]
                    worksheet[f"C{rank + 1}"] = f"{player_scores[pname]} + {score}"
                    player_scores[pname] += score
                    worksheet[f"A{rank + 1}"] = pname
                    worksheet[f"B{rank + 1}"] = player_scores[pname]

                image_path = filename.replace("-names", "")
                img = openpyxl.drawing.image.Image(image_path)
                img.anchor = 'D2'
                worksheet.add_image(img)

                screen_idx += 1


        # Update scores in output excel to serve as new input for the next ocr round
        sorted_players = {k: v for k, v in sorted(player_scores.items(), key=lambda item: item[1], reverse=True)}

        for idx, (pl_name, pl_score) in enumerate(sorted_players.items()):
            config_sheet[f"A{2+idx}"] = pl_name
            config_sheet[f"B{2+idx}"] = pl_score
        config_sheet["C2"] = nr_prev_games_parsed+screen_idx

        workbook.save(f'{params["output_sheet"]}')
        print(f'saved score results to {params["output_sheet"]}')

    if action == "tracks" or action == "run":

        workbook = openpyxl.load_workbook(f'{params["output_sheet"]}')
        sheet_nr = 1
        for root, _, files in os.walk(conf['frame_dir']+"/tracks"):
            files.sort(key=lambda f: int(re.sub('\D', '', f[:f.find('-')])))
            sheet_idx = 1
            last_trackname = ""
            for filename in files:
                raw_filename = filename
                if filename.find('-track-') > -1:
                    ext_idx = filename.index('.jpg')
                    trackname = filename[filename.index('-track-')+7:ext_idx]
                    trackname = base64.b64decode(trackname).decode("utf-8", "ignore")
                    if trackname == last_trackname:
                        continue

                    print(f"parse track {filename} : {trackname}  -> sheet {sheet_nr}")
                    last_trackname = trackname
                    worksheet = workbook.worksheets[sheet_nr]
                    worksheet[f"D1"] = trackname
                    print("add ", trackname)
                    sheet_nr += 1

        workbook.save(f'{params["output_sheet"]}')
        print(f'updated track for {params["output_sheet"]}')