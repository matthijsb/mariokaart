# pip install requirements.txt

# To test AABB coordinates & thresholds:
# python ocr.py --test_file test/frame1.jpg --aabb 1325,400,1750,800 --aabb_names 1400,410,1566,765

# To run:
# OPENCV_FFMPEG_READ_ATTEMPTS=100000 python ocr.py --movie movie.mp4 --sheet input.xlsx --output ./output --aabb 1325,400,1750,800 --aabb_names 1400,410,1566,765

import argparse
import time
import itertools
import cv2 as cv
import keras_ocr
import openpyxl

from thefuzz import process as fuzzy_match

from contextlib import contextmanager


# define the two output layer names for the EAST detector model that
# we are interested -- the first is the output probabilities and the
# second can be used to derive the bounding box coordinates of text
layerNames = [
    "feature_fusion/Conv_7/Sigmoid",
    "feature_fusion/concat_3"]


class Frame:
    def __init__(self, img, img_crop, img_names, img_scores, frame_number, video_frame_number, ts_second, match_count):
        self.img = img
        self.img_crop = img_crop
        self.img_names = img_names
        self.img_scores = img_scores
        self.frame_number = frame_number
        self.video_frame_number = video_frame_number
        self.ts_second = ts_second
        self.match_count = match_count


@contextmanager
def open_cv_video(filepath):
    cap = cv.VideoCapture(filepath)
    try:
        yield cap
    finally:
        cap.release()


def count_text_candidates(img, conf):
    #https://github.com/gifflet/opencv-text-detection/tree/master
    #https://github.com/songdejia/EAST

    #image = img[cropy[0]:cropy[1], cropx[0]:cropx[1]]
    image = img[conf["aabb_y1"]:conf["aabb_y2"], conf["aabb_x1"]:conf["aabb_x2"]]

    # set the new width and height and then determine the ratio in change
    # for both the width and height
    #(newW, newH) = (scaleW, scaleH)
    (newW, newH) = (conf["east_scale_w"], conf["east_scale_h"])

    # resize the image and grab the new image dimensions
    image = cv.resize(image, (newW, newH))
    (H, W) = image.shape[:2]

    net = conf["east_network"]

    # construct a blob from the image and then perform a forward pass of
    # the model to obtain the two output layer sets
    blob = cv.dnn.blobFromImage(image, 1.0, (W, H),(123.68, 116.78, 103.94), swapRB=True, crop=False)
    net.setInput(blob)
    (scores, geometry) = net.forward(layerNames)

    # grab the number of rows and columns from the scores volume, then
    # initialize our set of bounding box rectangles and corresponding
    # confidence scores
    (numRows, numCols) = scores.shape[2:4]
    confidences = []

    # loop over the number of rows
    for y in range(0, numRows):
        # extract the scores (probabilities), followed by the geometrical
        # data used to derive potential bounding box coordinates that
        # surround text
        scoresData = scores[0, 0, y]

        # loop over the number of columns
        for x in range(0, numCols):
            if scoresData[x] >= conf["east_confidence_threshold"]:
                confidences.append(scoresData[0])

    return image, len(confidences)


def analyze_frames(video_capture, conf):#sample_rate, east_threshold):
    candidate_frames = []
    video_frame_nr = -1
    capture_frame_nr = -1
    fast_forward = -1
    fps = int(video_capture.get(cv.CAP_PROP_FPS))

    while video_capture.isOpened():
        ret, frame = video_capture.read()
        if not ret:
            break

        video_frame_nr += 1

        # if video_frame_nr > 2000:
        #     break

        if fast_forward > -1:
            if video_frame_nr < fast_forward:
                continue
            else:
                fast_forward = -1

        if video_frame_nr / conf["movie_sample_rate"] != float(video_frame_nr // conf["movie_sample_rate"]):
           continue

        img_cropped, match_count = count_text_candidates(frame, conf)
        if match_count > conf["east_threshold"]:
            capture_frame_nr += 1
            img_cropped_names = frame[conf["aabb_names_y1"]:conf["aabb_names_y2"], conf["aabb_names_x1"]:conf["aabb_names_x2"]]
            #img_cropped_scores = frame[conf["aabb_scores_y1"]:conf["aabb_scores_y2"], conf["aabb_scores_x1"]:conf["aabb_scores_x2"]]
            img_cropped_scores = None
            candidate_frames.append(Frame(frame, img_cropped, img_cropped_names, img_cropped_scores, capture_frame_nr, video_frame_nr, video_frame_nr//fps, match_count))
            fast_forward = video_frame_nr + fps*10

            cv.imwrite(f"{conf['frame_dir']}/{capture_frame_nr}.jpg", frame)
            cv.imwrite(f"{conf['frame_dir']}/{capture_frame_nr}-aabb.jpg", img_cropped)
            cv.imwrite(f"{conf['frame_dir']}/{capture_frame_nr}-names.jpg", img_cropped_names)
            #cv.imwrite(f"{conf['frame_dir']}/{capture_frame_nr}-scores.jpg", img_cropped_scores)

            print("CANDIDATE FRAME FOUND: ",video_frame_nr,match_count, video_frame_nr / fps)

    print(f"ANALYZED {video_frame_nr} frames ({video_frame_nr//fps} seconds of video) with {len(candidate_frames)} candidates")
    return candidate_frames


ap = argparse.ArgumentParser()
ap.add_argument("--movie", required=True, help="the movie file to parse")
ap.add_argument("--sample_rate", required=False, default=100, help="only capture every X frame (defaults to every 100th frame)")
ap.add_argument("--sheet", required=True, help="the excel sheet to act as input (first sheet should contain player names on row A, scores on row B and NR sheets on row C1 (so we can continue scores in for stiched movie clips))")
ap.add_argument("--aabb", required=True, help="the dimensions of the aabb to check: x1,y1,x2,y2")
ap.add_argument("--aabb_names", required=True, help="the dimensions of the aabb to check for names: x1,y1,x2,y2")
ap.add_argument("--test_file", required=False, default=None, help="An image frame to test AABB coordinates & thresholds against")
ap.add_argument("--output", required=False, default=".", help="the directory for the output of the text detection pass")
ap.add_argument("--east_model", required=False, default="frozen_east_text_detection.pb", help="The frozen east text detection model to use")
ap.add_argument("--east_detect_width", required=False, default=320, help="Scale width factor for the EAST text detection (must be a multiple of 320?)")
ap.add_argument("--east_detect_height", required=False, default=320, help="Scale height factor for the EAST text detection (must be a multiple of 320?)")
ap.add_argument("--east_confidence_threshold", type=float, required=False, default=0.3, help="the threshold to use for the east confidence score")
ap.add_argument("--east_detection_threshold", type=float, required=False, default=100, help="the threshold to use for the expected nr of AABBS to detect for a frame with scores")
ap.add_argument("--fuzzy_threshold", required=False, default=100, help="The threshold to apply for accepting a detected text as a sure match")
ap.add_argument("--fuzzy_penalty", required=False, default=10, help="When we dont have a exact match we apply a penalty: (default 10) * every character wrongly matched")

params = vars(ap.parse_args())

aabb = [int(x) for x in params["aabb"].split(",")[0:4]]
aabb_names = [int(x) for x in params["aabb_names"].split(",")[0:4]]

conf = {
    "movie_file": params["movie"],
    "movie_sample_rate": params["sample_rate"],
    "frame_dir": params["output"],

    "east_model": params["east_model"],
    "east_scale_w": params["east_detect_width"],
    "east_scale_h": params["east_detect_height"],
    "east_confidence_threshold": params["east_confidence_threshold"],
    "east_threshold": params["east_detection_threshold"],     # for frames that contain a score list we

    "fuzzy_threshold": params["fuzzy_threshold"],
    "fuzzy_penalty": params["fuzzy_penalty"],

    "aabb_x1": aabb[0],
    "aabb_y1": aabb[1],
    "aabb_x2": aabb[2],
    "aabb_y2": aabb[3],

    "aabb_names_x1": aabb_names[0],
    "aabb_names_y1": aabb_names[1],
    "aabb_names_x2": aabb_names[2],
    "aabb_names_y2": aabb_names[3],
}

# load the pre-trained EAST text detector
conf["east_network"] = cv.dnn.readNet(conf["east_model"])


# Test a single frame to find thresholds
if params["test_file"]:
    frame = cv.imread(params["test_file"])
    #image = frame[aabb[1]:aabb[3], aabb[0]:aabb[2]]
    #cv.imwrite("testtt.jpg", image)
    img_cropped, match_count = count_text_candidates(frame, conf)
    print("Detection count (to set as guestimate): ", match_count)
    cv.imwrite("aabb.jpg", img_cropped)
    exit(1)

players = []
player_scores = {}
pos_scores = [13,11,9,8,7,6,5,4,3,2,1]

# Parse initial input (player names, scores

path = params["sheet"]
workbook = openpyxl.load_workbook(path)
#sheets = [x for x in workbook.sheetnames if x != "config"]
config_sheet = workbook["config"]#workbook.active
player_rec = config_sheet.cell(row = 2, column = 1)
score_rec = config_sheet.cell(row = 2, column = 2)
nr_prev_games_parsed = int(config_sheet.cell(row = 2, column = 3).value)
cx=2
while player_rec.value and cx < 100:
    player_name = player_rec.value.strip().lower()
    player_score = int(score_rec.value)
    print(f"Loaded {cx-1} - {player_name}: {player_score}")
    players.append(player_name)
    player_scores[player_name] = player_score
    cx += 1
    player_rec = config_sheet.cell(row=cx, column=1)
    score_rec = config_sheet.cell(row=cx, column=2)
print("\n")

#players = [xx.strip() for xx in params["players"].split()]
#players = [xx.strip() for xx in params["players"].split()]
#player_scores = {x:conf["start_score"].get(x, 0) for x in players}

# As a preprocessing step, we jump through frames and apply text detection using the EAST text model
start = time.time()
with open_cv_video(conf["movie_file"]) as cap:
    candidate_frames = analyze_frames(cap, conf)#sample_rate, 100)
end = time.time()
print("\nCandidate sampling took {:.2f} seconds\n".format(end - start))
frames = [x.img_names for x in candidate_frames]
#frames = [cv.imread('./0-names.jpg'),]

#frames = [cv.imread('./0-names.jpg'),cv.imread('./1-names.jpg'),cv.imread('./2-names.jpg'),cv.imread('./3-names.jpg'),cv.imread('./4-names.jpg')]

# Then we apply text recognition on these candidate images
#https://keras-ocr.readthedocs.io/en/latest/
pipeline = keras_ocr.pipeline.Pipeline()

# TODO:
# Keras-OCR detect does not recognize digits
# keras trainen op digits?: https://stackoverflow.com/questions/72433580/how-make-keras-ocr-default-model-recognize-only-numbers
# https://towardsdatascience.com/build-a-multi-digit-detector-with-keras-and-opencv-b97e3cd3b37

# We add players with a space in their name, since keras will split on space
multi_name_list = [x for x in players if (" " in x) == True]
multi_name = {}
for sp in multi_name_list:
    multi_name[sp] = list(itertools.permutations(sp.split()))

prediction_groups = pipeline.recognize(frames)
for idx, prediction_result in enumerate(prediction_groups):
    print(f"\n\nRONDE {idx+1}")

    ocr_results = [x[0] for x in prediction_result]
    print(f"OCR results: {ocr_results}")
    if len(ocr_results) < len(players):
        raise Exception(f"ERROR: OCR results < PLAYER COUNT")

    possible_names = set(players)
    ocr_indices_resolved = set()
    match_iter = 0
    pos_skip = 0
    multi_idx = {}
    skip_next = False
    match_fuz = {}
    nr_players_resolved = 0

    while nr_players_resolved < len(players) and match_iter < 2:
        match_iter += 1
        print(f"\nMatching iteration {match_iter}")

        # After the first iteration, narrow the list of fuzze matches we can make
        for t in match_fuz.keys():
            if match_fuz[t]:
                if match_fuz[t]["name"] in possible_names:
                    possible_names.remove(match_fuz[t]["name"])
                ocr_indices_resolved.add(match_fuz[t]["pos"])

        # Loop through all ocr
        for x in range(len(ocr_results)):

            if skip_next:
                pos_skip += 1
                skip_next = False
                continue

            if x in ocr_indices_resolved:
                continue

            # TODO: support spaced names with more than 1 space in the name
            spaced_names = []
            if (x + 1) < len(ocr_results) - 1:
                spaced_names = [f"{ocr_results[x]} {ocr_results[x+1]}"]

            fuzzy_results = []
            fuzzy_results += fuzzy_match.extract(ocr_results[x], list(possible_names))
            print(f"Matching: {x} - {ocr_results[x]} - {list(possible_names)}")

            matched_name = fuzzy_results[0][0]

            score = fuzzy_results[0][1]
            delta = abs(len(ocr_results[x]) - len(matched_name))
            penalty = 0
            if score < conf["fuzzy_threshold"]:
                # TODO: maybe add penalty on phonetic difference: https://medium.com/data-science-in-your-pocket/phonetics-based-fuzzy-string-matching-algorithms-8399aea04718
                penalty = delta * conf["fuzzy_penalty"] #len((set(matched_name)).intersection(set(ocr_results[x]))) * 10
                score -= penalty # apply a higher penalty to length mismatch then to char mismatches

            if match_fuz.get(matched_name, False):
                if match_fuz[matched_name]["score"] >= score:
                    print(f"skippie {matched_name} {match_fuz[matched_name]['score']} > {score}")
                    continue

            ocr_match = ocr_results[x]
            if matched_name in multi_name:
                if f"{ocr_results[x]} {ocr_results[x+1]}" in multi_name:
                    print(f"FOUND A PERFECT MULTI NAME MATCH {ocr_results[x]} {ocr_results[x+1]}: {score}")
                    score = 100
                    ocr_indices_resolved.add(x)
                    ocr_indices_resolved.add(x+1) # TODO: in case we have names with multiple spaces, we should add them all here
                else:
                    print(f"FOUND A FUZZY MULTI NAME MATCH {ocr_results[x]} {ocr_results[x + 1]}: {score}")

                ocr_match = f"{x>0 and ocr_results[x-1] or '-'} {ocr_results[x]} {ocr_results[x+1]}"
                skip_next = True    # TODO: in case we have names with multiple spaces, we should use a counter on nr spaces here instead

            # Remove all earlier possible matches for this name that we're not sure of
            for k in match_fuz.keys():
                if match_fuz.get(k, False) and k == matched_name and match_fuz[k]["score"] < score:
                    print(f"Delete {matched_name} - {x} - {match_fuz[k]['score']} < {score} - ocr: {x} {ocr_results[x]} - matches: {fuzzy_results}")
                    nr_players_resolved -= 1
                    match_fuz[k] = False

            match_fuz[matched_name] = {
                "name": matched_name,
                "score": score,
                "pos": x,
                "delta": delta,
                "ocr": ocr_match,
            }
            print(f"Add {matched_name} - {x} - score: {score} - ocr: {x} {ocr_results[x]} - matches: {fuzzy_results}")

            if score == 100:
                possible_names.remove(matched_name)
                ocr_indices_resolved.add(x)

            nr_players_resolved += 1

    # preprocess sorting value
    score_list = []
    pos_correction = 1
    for t in match_fuz.keys():
        if match_fuz[t]:
            match_fuz[t]["pos"] += pos_correction
            score_list.append(match_fuz[t])
            if match_fuz[t]["name"] in multi_name:
                pos_correction -= 1 # TODO: account for space_count in case name consists > 1 space

    worksheet = workbook.create_sheet(f"Ronde {nr_prev_games_parsed+idx+1}", nr_prev_games_parsed+idx+1)
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 10

    rankings = sorted(score_list, key=lambda d: d['pos'])

    if len(rankings) < len(players):
        for rank in rankings:
            pname = rank["name"]
            print(rank["pos"], pname, player_scores[pname])#, pos_scores[rank["pos"] - 1], player_scores[pname] + pos_scores[rank["pos"] - 1])

        raise Exception(f"NOT ALL PLAYERS HAVE BEEN MATCHED {len(rankings)} < {len(players)}")


    for rx, rank in enumerate(rankings):
        pname = rank["name"]
        score = pos_scores[rank["pos"]-1]
        player_scores[pname] += score
        worksheet[f"A{rx + 1}"] = pname
        worksheet[f"B{rx + 1}"] = player_scores[pname]

        print(rank["pos"], pname, player_scores[pname], pos_scores[rank["pos"] - 1], player_scores[pname] + pos_scores[rank["pos"] - 1])

    img = openpyxl.drawing.image.Image(f"{conf['frame_dir']}/{idx}-aabb.jpg")
    img.anchor = 'D1'
    worksheet.add_image(img)

config_sheet["C2"] = nr_prev_games_parsed+idx+1
workbook.save(f'{params["output"]}/output.xlsx')
